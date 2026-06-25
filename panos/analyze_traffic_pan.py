import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Any, Union
from dataclasses import dataclass, asdict
from enum import Enum
import json
import re
import os
from pathlib import Path

# Excel export libraries
try:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.chart import BarChart, Reference, PieChart, LineChart
    from openpyxl.chart.label import DataLabelList
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("Warning: openpyxl not installed. Install with: pip install openpyxl")


class LogType(Enum):
    """Enum for different PAN-OS log types"""
    TRAFFIC = "traffic"
    URL = "url"
    THREAT = "threat"
    SYSTEM = "system"
    CONFIG = "config"


class SessionEndReason(Enum):
    """Session end reasons for traffic logs"""
    NORMAL = "normal"
    TIMEOUT = "timeout"
    DENIED = "denied"
    RESET = "reset"
    AGED_OUT = "aged-out"
    UNKNOWN = "unknown"


class PanOSLogAnalyzer:
    """
    Comprehensive PAN-OS log analyzer using pandas for efficient data processing
    and Excel report generation
    """
    
    def __init__(self):
        self.traffic_df = pd.DataFrame()
        self.url_df = pd.DataFrame()
        self.threat_df = pd.DataFrame()
        self._setup_column_mappings()
    
    def _setup_column_mappings(self):
        """Define column mappings for different log types"""
        # Common column mappings for traffic logs
        self.traffic_mapping = {
            'Time': 'timestamp',
            'Serial': 'serial_number',
            'Seq': 'seq_no',
            'Source IP': 'source_ip',
            'Dest IP': 'dest_ip',
            'Source Port': 'source_port',
            'Dest Port': 'dest_port',
            'Protocol': 'protocol',
            'Application': 'application',
            'Action': 'action',
            'Bytes Sent': 'bytes_sent',
            'Bytes Received': 'bytes_received',
            'Packets Sent': 'packets_sent',
            'Packets Received': 'packets_received',
            'Session End Reason': 'session_end_reason',
            'Duration': 'duration',
            'Source Zone': 'source_zone',
            'Dest Zone': 'dest_zone',
            'NAT Source IP': 'nat_source_ip',
            'NAT Dest IP': 'nat_dest_ip',
            'Rule Name': 'rule_name',
            'Category': 'category',
            'Source Country': 'source_country',
            'Dest Country': 'dest_country'
        }
        
        # Column mappings for URL logs
        self.url_mapping = {
            'Time': 'timestamp',
            'Serial': 'serial_number',
            'Seq': 'seq_no',
            'Source IP': 'source_ip',
            'Dest IP': 'dest_ip',
            'Source Port': 'source_port',
            'Dest Port': 'dest_port',
            'URL': 'url',
            'Category': 'category',
            'Action': 'action',
            'Threat ID': 'threat_id',
            'Threat Name': 'threat_name',
            'Source User': 'source_user',
            'Source Zone': 'source_zone',
            'Dest Zone': 'dest_zone',
            'Rule Name': 'rule_name',
            'Content Type': 'content_type'
        }
    
    def load_traffic_logs(self, filepath: str, **kwargs) -> pd.DataFrame:
        """
        Load traffic logs from CSV file using pandas
        
        Args:
            filepath: Path to CSV file
            **kwargs: Additional pandas read_csv arguments
        """
        try:
            # Auto-detect delimiter
            with open(filepath, 'r', encoding='utf-8') as f:
                sample = f.read(1024)
                delimiter = self._detect_delimiter(sample)
            
            # Read CSV with pandas
            df = pd.read_csv(filepath, delimiter=delimiter, **kwargs)
            
            # Rename columns
            df = df.rename(columns=self.traffic_mapping)
            
            # Convert timestamp
            if 'timestamp' in df.columns:
                df['timestamp'] = pd.to_datetime(df['timestamp'], errors='coerce')
            
            # Convert numeric columns
            numeric_columns = ['source_port', 'dest_port', 'bytes_sent', 'bytes_received', 
                             'packets_sent', 'packets_received', 'seq_no']
            for col in numeric_columns:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors='coerce')
            
            # Calculate total bytes and packets
            if 'bytes_sent' in df.columns and 'bytes_received' in df.columns:
                df['total_bytes'] = df['bytes_sent'] + df['bytes_received']
                df['total_packets'] = df.get('packets_sent', 0) + df.get('packets_received', 0)
            
            # Store in class
            self.traffic_df = df
            
            # Add additional derived columns
            self._add_derived_columns('traffic')
            
            print(f"✅ Loaded {len(df)} traffic logs from {filepath}")
            return df
            
        except Exception as e:
            print(f"❌ Error loading traffic logs: {e}")
            return pd.DataFrame()
    
    def load_url_logs(self, filepath: str, **kwargs) -> pd.DataFrame:
        """
        Load URL logs from CSV file using pandas
        """
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                sample = f.read(1024)
                delimiter = self._detect_delimiter(sample)
            
            df = pd.read_csv(filepath, delimiter=delimiter, **kwargs)
            df = df.rename(columns=self.url_mapping)
            
            if 'timestamp' in df.columns:
                df['timestamp'] = pd.to_datetime(df['timestamp'], errors='coerce')
            
            numeric_columns = ['source_port', 'dest_port', 'threat_id', 'seq_no']
            for col in numeric_columns:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors='coerce')
            
            # Parse URL domain
            if 'url' in df.columns:
                df['domain'] = df['url'].apply(self._extract_domain)
            
            self.url_df = df
            print(f"✅ Loaded {len(df)} URL logs from {filepath}")
            return df
            
        except Exception as e:
            print(f"❌ Error loading URL logs: {e}")
            return pd.DataFrame()
    
    def _detect_delimiter(self, sample: str) -> str:
        """Auto-detect CSV delimiter"""
        if '\t' in sample:
            return '\t'
        elif ';' in sample:
            return ';'
        else:
            return ','
    
    def _extract_domain(self, url: str) -> str:
        """Extract domain from URL"""
        if pd.isna(url) or not url:
            return 'unknown'
        
        try:
            # Remove protocol
            url = re.sub(r'^https?://', '', url)
            # Get domain (first part before / or ?)
            domain = url.split('/')[0]
            # Remove port if present
            domain = domain.split(':')[0]
            return domain.lower()
        except:
            return 'unknown'
    
    def _add_derived_columns(self, log_type: str):
        """Add derived columns to dataframes"""
        if log_type == 'traffic' and not self.traffic_df.empty:
            df = self.traffic_df
            
            # Time-based columns
            if 'timestamp' in df.columns:
                df['hour'] = df['timestamp'].dt.hour
                df['day_of_week'] = df['timestamp'].dt.day_name()
                df['date'] = df['timestamp'].dt.date
                df['month'] = df['timestamp'].dt.month_name()
            
            # Traffic direction
            if 'source_ip' in df.columns and 'dest_ip' in df.columns:
                df['is_internal_to_internal'] = df['source_ip'].apply(self._is_private_ip) & df['dest_ip'].apply(self._is_private_ip)
                df['is_internal_to_external'] = df['source_ip'].apply(self._is_private_ip) & ~df['dest_ip'].apply(self._is_private_ip)
                df['is_external_to_internal'] = ~df['source_ip'].apply(self._is_private_ip) & df['dest_ip'].apply(self._is_private_ip)
    
    def _is_private_ip(self, ip: str) -> bool:
        """Check if IP is private"""
        if pd.isna(ip) or not ip:
            return False
        
        # Check if it's a string
        if not isinstance(ip, str):
            return False
            
        # Private IP ranges
        private_ranges = [
            (r'^10\.', '10.0.0.0/8'),
            (r'^172\.(1[6-9]|2[0-9]|3[0-1])\.', '172.16.0.0/12'),
            (r'^192\.168\.', '192.168.0.0/16'),
            (r'^127\.', '127.0.0.0/8')
        ]
        
        for pattern, _ in private_ranges:
            if re.match(pattern, ip):
                return True
        return False
    
    def generate_traffic_report(self) -> Dict[str, Any]:
        """Generate comprehensive traffic report statistics"""
        if self.traffic_df.empty:
            return {'error': 'No traffic logs loaded'}
        
        df = self.traffic_df
        
        # Basic statistics
        report = {
            'total_logs': len(df),
            'total_bytes': df['total_bytes'].sum() if 'total_bytes' in df.columns else 0,
            'total_packets': df['total_packets'].sum() if 'total_packets' in df.columns else 0,
            'unique_ips': len(pd.concat([df['source_ip'], df['dest_ip']]).unique()),
            'unique_applications': df['application'].nunique() if 'application' in df.columns else 0,
            'unique_sessions': len(df['seq_no'].unique()) if 'seq_no' in df.columns else 0,
        }
        
        # Top applications
        if 'application' in df.columns:
            report['top_applications'] = df['application'].value_counts().head(10).to_dict()
            report['top_applications_by_bytes'] = df.groupby('application')['total_bytes'].sum().sort_values(ascending=False).head(10).to_dict()
        
        # Top source and destination IPs
        if 'source_ip' in df.columns:
            report['top_source_ips'] = df['source_ip'].value_counts().head(10).to_dict()
        if 'dest_ip' in df.columns:
            report['top_dest_ips'] = df['dest_ip'].value_counts().head(10).to_dict()
        
        # Session end reasons
        if 'session_end_reason' in df.columns:
            report['session_end_reasons'] = df['session_end_reason'].value_counts().to_dict()
        
        # Actions
        if 'action' in df.columns:
            report['actions'] = df['action'].value_counts().to_dict()
        
        # Protocols
        if 'protocol' in df.columns:
            report['protocols'] = df['protocol'].value_counts().to_dict()
        
        # Time-based statistics
        if 'hour' in df.columns:
            report['traffic_by_hour'] = df.groupby('hour').size().to_dict()
        if 'day_of_week' in df.columns:
            report['traffic_by_day'] = df.groupby('day_of_week').size().to_dict()
        
        # Traffic direction
        if 'is_internal_to_internal' in df.columns:
            report['internal_to_internal'] = df['is_internal_to_internal'].sum()
            report['internal_to_external'] = df['is_internal_to_external'].sum()
            report['external_to_internal'] = df['is_external_to_internal'].sum()
        
        return report
    
    def generate_url_report(self) -> Dict[str, Any]:
        """Generate comprehensive URL report statistics"""
        if self.url_df.empty:
            return {'error': 'No URL logs loaded'}
        
        df = self.url_df
        
        report = {
            'total_logs': len(df),
            'unique_urls': df['url'].nunique() if 'url' in df.columns else 0,
            'unique_categories': df['category'].nunique() if 'category' in df.columns else 0,
            'unique_domains': df['domain'].nunique() if 'domain' in df.columns else 0,
        }
        
        # Categories
        if 'category' in df.columns:
            report['categories'] = df['category'].value_counts().to_dict()
        
        # Actions
        if 'action' in df.columns:
            report['actions'] = df['action'].value_counts().to_dict()
            report['blocked_count'] = df[df['action'].str.lower().str.contains('block|deny', na=False)].shape[0]
            report['allowed_count'] = df[df['action'].str.lower().str.contains('allow', na=False)].shape[0]
        
        # Top domains
        if 'domain' in df.columns:
            report['top_domains'] = df['domain'].value_counts().head(10).to_dict()
        
        # Top URL categories by count
        if 'category' in df.columns:
            report['top_categories'] = df['category'].value_counts().head(10).to_dict()
        
        # Time-based statistics
        if 'hour' in df.columns:
            report['urls_by_hour'] = df.groupby('hour').size().to_dict()
        if 'day_of_week' in df.columns:
            report['urls_by_day'] = df.groupby('day_of_week').size().to_dict()
        
        # Threat information
        if 'threat_id' in df.columns:
            report['unique_threats'] = df['threat_id'].nunique()
            report['top_threats'] = df['threat_name'].value_counts().head(10).to_dict() if 'threat_name' in df.columns else {}
        
        return report
    
    def generate_excel_report(self, output_path: str = 'panos_report.xlsx'):
        """
        Generate comprehensive Excel report with multiple sheets and charts
        
        Args:
            output_path: Path where the Excel file will be saved
        """
        if not OPENPYXL_AVAILABLE:
            print("❌ openpyxl is not installed. Install with: pip install openpyxl")
            return
        
        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # Write data sheets
                self._write_traffic_sheets(writer)
                self._write_url_sheets(writer)
                self._write_summary_sheets(writer)
                self._write_analysis_sheets(writer)
                
                # Apply formatting
                self._apply_excel_formatting(writer)
                
            print(f"✅ Excel report generated: {output_path}")
            
        except Exception as e:
            print(f"❌ Error generating Excel report: {e}")
    
    def _write_traffic_sheets(self, writer: pd.ExcelWriter):
        """Write traffic log sheets"""
        if not self.traffic_df.empty:
            # Main traffic data
            traffic_display = self.traffic_df.copy()
            # Select key columns for display
            display_cols = ['timestamp', 'source_ip', 'dest_ip', 'application', 'action', 
                           'total_bytes', 'protocol', 'source_port', 'dest_port']
            display_cols = [col for col in display_cols if col in traffic_display.columns]
            if display_cols:
                traffic_display[display_cols].to_excel(writer, sheet_name='Traffic Logs', index=False)
            
            # Traffic summary
            summary = self._create_traffic_summary_df()
            if not summary.empty:
                summary.to_excel(writer, sheet_name='Traffic Summary', index=False)
            
            # Top applications
            if 'application' in self.traffic_df.columns:
                top_apps = self.traffic_df['application'].value_counts().reset_index()
                top_apps.columns = ['Application', 'Count']
                top_apps.to_excel(writer, sheet_name='Top Applications', index=False)
            
            # Traffic by hour
            if 'hour' in self.traffic_df.columns:
                hourly = self.traffic_df.groupby('hour').size().reset_index()
                hourly.columns = ['Hour', 'Count']
                hourly.to_excel(writer, sheet_name='Traffic by Hour', index=False)
    
    def _write_url_sheets(self, writer: pd.ExcelWriter):
        """Write URL log sheets"""
        if not self.url_df.empty:
            # Main URL data
            url_display = self.url_df.copy()
            display_cols = ['timestamp', 'source_ip', 'dest_ip', 'url', 'category', 
                           'action', 'domain', 'threat_name']
            display_cols = [col for col in display_cols if col in url_display.columns]
            if display_cols:
                url_display[display_cols].to_excel(writer, sheet_name='URL Logs', index=False)
            
            # URL summary
            summary = self._create_url_summary_df()
            if not summary.empty:
                summary.to_excel(writer, sheet_name='URL Summary', index=False)
            
            # Categories
            if 'category' in self.url_df.columns:
                categories = self.url_df['category'].value_counts().reset_index()
                categories.columns = ['Category', 'Count']
                categories.to_excel(writer, sheet_name='URL Categories', index=False)
    
    def _write_summary_sheets(self, writer: pd.ExcelWriter):
        """Write overall summary sheets"""
        # Overall dashboard
        dashboard_data = self._create_dashboard_df()
        if not dashboard_data.empty:
            dashboard_data.to_excel(writer, sheet_name='Dashboard', index=False)
        
        # Combined statistics
        stats_data = self._create_statistics_df()
        if not stats_data.empty:
            stats_data.to_excel(writer, sheet_name='Statistics', index=False)
    
    def _write_analysis_sheets(self, writer: pd.ExcelWriter):
        """Write analysis sheets"""
        # Traffic analysis with Pivot Tables
        if not self.traffic_df.empty:
            self._create_pivot_sheets(writer)
        
        # Time series analysis
        if not self.traffic_df.empty and 'timestamp' in self.traffic_df.columns:
            self._create_timeseries_sheets(writer)
    
    def _create_traffic_summary_df(self) -> pd.DataFrame:
        """Create traffic summary dataframe"""
        if self.traffic_df.empty:
            return pd.DataFrame()
        
        df = self.traffic_df
        summary = {
            'Metric': ['Total Logs', 'Total Bytes (GB)', 'Total Packets', 'Unique IPs', 
                      'Unique Applications', 'Average Bytes per Session'],
            'Value': [
                len(df),
                df['total_bytes'].sum() / (1024**3) if 'total_bytes' in df.columns else 0,
                df['total_packets'].sum() if 'total_packets' in df.columns else 0,
                len(pd.concat([df['source_ip'], df['dest_ip']]).unique()),
                df['application'].nunique() if 'application' in df.columns else 0,
                df['total_bytes'].mean() if 'total_bytes' in df.columns else 0
            ]
        }
        return pd.DataFrame(summary)
    
    def _create_url_summary_df(self) -> pd.DataFrame:
        """Create URL summary dataframe"""
        if self.url_df.empty:
            return pd.DataFrame()
        
        df = self.url_df
        summary = {
            'Metric': ['Total Logs', 'Unique URLs', 'Unique Categories', 'Unique Domains',
                      'Blocked URLs', 'Allowed URLs'],
            'Value': [
                len(df),
                df['url'].nunique() if 'url' in df.columns else 0,
                df['category'].nunique() if 'category' in df.columns else 0,
                df['domain'].nunique() if 'domain' in df.columns else 0,
                df[df['action'].str.lower().str.contains('block|deny', na=False)].shape[0] if 'action' in df.columns else 0,
                df[df['action'].str.lower().str.contains('allow', na=False)].shape[0] if 'action' in df.columns else 0
            ]
        }
        return pd.DataFrame(summary)
    
    def _create_dashboard_df(self) -> pd.DataFrame:
        """Create overall dashboard data"""
        dashboard = {
            'Traffic Logs': [len(self.traffic_df) if not self.traffic_df.empty else 0],
            'URL Logs': [len(self.url_df) if not self.url_df.empty else 0],
            'Total Logs': [(len(self.traffic_df) if not self.traffic_df.empty else 0) + 
                          (len(self.url_df) if not self.url_df.empty else 0)]
        }
        
        # Add traffic metrics
        if not self.traffic_df.empty:
            dashboard['Total Traffic (GB)'] = [self.traffic_df['total_bytes'].sum() / (1024**3) if 'total_bytes' in self.traffic_df.columns else 0]
            dashboard['Unique Applications'] = [self.traffic_df['application'].nunique() if 'application' in self.traffic_df.columns else 0]
        
        # Add URL metrics
        if not self.url_df.empty:
            dashboard['URL Categories'] = [self.url_df['category'].nunique() if 'category' in self.url_df.columns else 0]
            dashboard['Unique Domains'] = [self.url_df['domain'].nunique() if 'domain' in self.url_df.columns else 0]
        
        df = pd.DataFrame(dashboard)
        return df.T.reset_index()
    
    def _create_statistics_df(self) -> pd.DataFrame:
        """Create comprehensive statistics dataframe"""
        stats = []
        
        # Traffic stats
        if not self.traffic_df.empty:
            df = self.traffic_df
            stats.append(['Traffic Logs Count', len(df)])
            if 'total_bytes' in df.columns:
                stats.append(['Traffic Total Bytes', f"{df['total_bytes'].sum():,.0f}"])
                stats.append(['Traffic Average Bytes', f"{df['total_bytes'].mean():,.0f}"])
                stats.append(['Traffic Max Bytes', f"{df['total_bytes'].max():,.0f}"])
            if 'timestamp' in df.columns:
                stats.append(['Traffic Time Range Start', df['timestamp'].min()])
                stats.append(['Traffic Time Range End', df['timestamp'].max()])
        
        # URL stats
        if not self.url_df.empty:
            df = self.url_df
            stats.append(['URL Logs Count', len(df)])
            if 'category' in df.columns:
                stats.append(['Unique URL Categories', df['category'].nunique()])
            if 'timestamp' in df.columns:
                stats.append(['URL Time Range Start', df['timestamp'].min()])
                stats.append(['URL Time Range End', df['timestamp'].max()])
        
        return pd.DataFrame(stats, columns=['Metric', 'Value'])
    
    def _create_pivot_sheets(self, writer: pd.ExcelWriter):
        """Create pivot table sheets"""
        df = self.traffic_df
        
        # Applications by action
        if 'application' in df.columns and 'action' in df.columns:
            pivot = pd.crosstab(df['application'], df['action'])
            pivot.to_excel(writer, sheet_name='Apps vs Actions')
        
        # Source IPs by action
        if 'source_ip' in df.columns and 'action' in df.columns:
            top_ips = df['source_ip'].value_counts().head(20).index
            pivot = pd.crosstab(df[df['source_ip'].isin(top_ips)]['source_ip'], 
                              df[df['source_ip'].isin(top_ips)]['action'])
            if not pivot.empty:
                pivot.to_excel(writer, sheet_name='Top IPs vs Actions')
    
    def _create_timeseries_sheets(self, writer: pd.ExcelWriter):
        """Create time series analysis sheets"""
        df = self.traffic_df
        
        # Daily traffic
        if 'date' in df.columns:
            daily = df.groupby('date').agg({
                'total_bytes': 'sum',
                'source_ip': 'count'
            }).rename(columns={'source_ip': 'count'})
            daily.to_excel(writer, sheet_name='Daily Traffic')
        
        # Hourly traffic pattern
        if 'hour' in df.columns:
            hourly_stats = df.groupby('hour').agg({
                'total_bytes': ['mean', 'sum'],
                'source_ip': 'count'
            }).round(2)
            hourly_stats.to_excel(writer, sheet_name='Hourly Patterns')
    
    def _apply_excel_formatting(self, writer: pd.ExcelWriter):
        """Apply formatting to Excel sheets"""
        workbook = writer.book
        
        # Define styles
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Apply to all sheets
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            
            # Format headers
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def create_charts(self, output_dir: str = 'charts'):
        """
        Create and save charts using matplotlib
        
        Args:
            output_dir: Directory to save charts
        """
        try:
            import matplotlib.pyplot as plt
            import seaborn as sns
            
            # Create output directory
            Path(output_dir).mkdir(parents=True, exist_ok=True)
            
            # Traffic charts
            if not self.traffic_df.empty:
                self._create_traffic_charts(output_dir, plt, sns)
            
            # URL charts
            if not self.url_df.empty:
                self._create_url_charts(output_dir, plt, sns)
            
            print(f"✅ Charts saved to {output_dir}/")
            
        except ImportError:
            print("⚠️ Matplotlib/seaborn not installed. Skipping chart creation.")
            print("Install with: pip install matplotlib seaborn")
    
    def _create_traffic_charts(self, output_dir: str, plt, sns):
        """Create traffic charts"""
        df = self.traffic_df
        
        # Set style
        sns.set_style("whitegrid")
        
        # 1. Top Applications (Bar Chart)
        if 'application' in df.columns:
            fig, ax = plt.subplots(figsize=(12, 6))
            top_apps = df['application'].value_counts().head(10)
            top_apps.plot(kind='bar', ax=ax, color='skyblue')
            ax.set_title('Top 10 Applications', fontsize=14, fontweight='bold')
            ax.set_xlabel('Application')
            ax.set_ylabel('Number of Sessions')
            plt.xticks(rotation=45, ha='right')
            plt.tight_layout()
            plt.savefig(f'{output_dir}/top_applications.png', dpi=300, bbox_inches='tight')
            plt.close()
        
        # 2. Traffic by Hour (Line Chart)
        if 'hour' in df.columns:
            fig, ax = plt.subplots(figsize=(12, 6))
            hourly = df.groupby('hour').size()
            hourly.plot(kind='line', marker='o', ax=ax, color='green')
            ax.set_title('Traffic Distribution by Hour', fontsize=14, fontweight='bold')
            ax.set_xlabel('Hour of Day')
            ax.set_ylabel('Number of Sessions')
            ax.grid(True, alpha=0.3)
            plt.tight_layout()
            plt.savefig(f'{output_dir}/traffic_by_hour.png', dpi=300, bbox_inches='tight')
            plt.close()
        
        # 3. Session End Reasons (Pie Chart)
        if 'session_end_reason' in df.columns:
            fig, ax = plt.subplots(figsize=(10, 8))
            reasons = df['session_end_reason'].value_counts()
            reasons.plot(kind='pie', autopct='%1.1f%%', ax=ax, startangle=90)
            ax.set_title('Session End Reasons', fontsize=14, fontweight='bold')
            ax.set_ylabel('')
            plt.tight_layout()
            plt.savefig(f'{output_dir}/session_end_reasons.png', dpi=300, bbox_inches='tight')
            plt.close()
        
        # 4. Traffic Direction (Donut Chart)
        if 'is_internal_to_internal' in df.columns:
            fig, ax = plt.subplots(figsize=(10, 8))
            directions = {
                'Internal→Internal': df['is_internal_to_internal'].sum(),
                'Internal→External': df['is_internal_to_external'].sum(),
                'External→Internal': df['is_external_to_internal'].sum()
            }
            values = list(directions.values())
            labels = list(directions.keys())
            colors = ['#66b3ff', '#ff9999', '#99ff99']
            
            # Create donut chart
            wedges, texts, autotexts = ax.pie(values, labels=labels, autopct='%1.1f%%',
                                              colors=colors, startangle=90)
            # Draw circle for donut
            centre_circle = plt.Circle((0, 0), 0.70, fc='white')
            fig.gca().add_artist(centre_circle)
            ax.set_title('Traffic Flow Direction', fontsize=14, fontweight='bold')
            plt.tight_layout()
            plt.savefig(f'{output_dir}/traffic_direction.png', dpi=300, bbox_inches='tight')
            plt.close()
    
    def _create_url_charts(self, output_dir: str, plt, sns):
        """Create URL charts"""
        df = self.url_df
        
        # 1. URL Categories (Bar Chart)
        if 'category' in df.columns:
            fig, ax = plt.subplots(figsize=(12, 6))
            top_cats = df['category'].value_counts().head(10)
            top_cats.plot(kind='bar', ax=ax, color='orange')
            ax.set_title('Top 10 URL Categories', fontsize=14, fontweight='bold')
            ax.set_xlabel('Category')
            ax.set_ylabel('Count')
            plt.xticks(rotation=45, ha='right')
            plt.tight_layout()
            plt.savefig(f'{output_dir}/url_categories.png', dpi=300, bbox_inches='tight')
            plt.close()
        
        # 2. URL Actions (Pie Chart)
        if 'action' in df.columns:
            fig, ax = plt.subplots(figsize=(10, 8))
            actions = df['action'].value_counts()
            actions.plot(kind='pie', autopct='%1.1f%%', ax=ax, startangle=90,
                        colors=['green', 'red', 'blue', 'orange'])
            ax.set_title('URL Filtering Actions', fontsize=14, fontweight='bold')
            ax.set_ylabel('')
            plt.tight_layout()
            plt.savefig(f'{output_dir}/url_actions.png', dpi=300, bbox_inches='tight')
            plt.close()
        
        # 3. Top Domains (Horizontal Bar Chart)
        if 'domain' in df.columns:
            fig, ax = plt.subplots(figsize=(12, 6))
            top_domains = df['domain'].value_counts().head(10)
            top_domains.plot(kind='barh', ax=ax, color='purple')
            ax.set_title('Top 10 Domains', fontsize=14, fontweight='bold')
            ax.set_xlabel('Count')
            ax.set_ylabel('Domain')
            plt.tight_layout()
            plt.savefig(f'{output_dir}/top_domains.png', dpi=300, bbox_inches='tight')
            plt.close()
    
    def analyze_traffic_patterns(self) -> Dict[str, Any]:
        """Advanced traffic pattern analysis"""
        if self.traffic_df.empty:
            return {'error': 'No traffic logs loaded'}
        
        df = self.traffic_df
        patterns = {}
        
        # Peak hour analysis
        if 'hour' in df.columns:
            hour_counts = df.groupby('hour').size()
            peak_hour = hour_counts.idxmax() if not hour_counts.empty else None
            peak_count = hour_counts.max() if not hour_counts.empty else 0
            patterns['peak_hour'] = {'hour': peak_hour, 'sessions': peak_count}
            
            # Off-peak hours (lowest traffic)
            off_peak_hour = hour_counts.idxmin() if not hour_counts.empty else None
            patterns['off_peak_hour'] = {'hour': off_peak_hour, 'sessions': hour_counts.min() if not hour_counts.empty else 0}
        
        # Application usage patterns
        if 'application' in df.columns:
            app_counts = df['application'].value_counts()
            top_apps = app_counts.head(5)
            patterns['top_applications'] = top_apps.to_dict()
            patterns['application_diversity'] = len(app_counts)
        
        # Traffic volume over time
        if 'date' in df.columns:
            daily_traffic = df.groupby('date').size()
            if len(daily_traffic) > 1:
                patterns['daily_avg'] = daily_traffic.mean()
                patterns['daily_std'] = daily_traffic.std()
                patterns['max_daily_traffic'] = daily_traffic.max()
                patterns['min_daily_traffic'] = daily_traffic.min()
        
        # Protocol distribution
        if 'protocol' in df.columns:
            patterns['protocols'] = df['protocol'].value_counts().to_dict()
        
        return patterns
    
    def detect_anomalies(self, method: str = 'zscore') -> pd.DataFrame:
        """
        Detect anomalies in traffic patterns
        
        Args:
            method: Detection method ('zscore' or 'iqr')
        """
        if self.traffic_df.empty:
            return pd.DataFrame()
        
        df = self.traffic_df.copy()
        
        if 'total_bytes' not in df.columns:
            return pd.DataFrame()
        
        anomalies = []
        
        if method == 'zscore':
            # Z-score method
            mean_bytes = df['total_bytes'].mean()
            std_bytes = df['total_bytes'].std()
            
            if std_bytes > 0:
                df['zscore'] = (df['total_bytes'] - mean_bytes) / std_bytes
                df['is_anomaly'] = abs(df['zscore']) > 3  # 3 standard deviations
                anomalies = df[df['is_anomaly']].copy()
        
        elif method == 'iqr':
            # IQR method
            Q1 = df['total_bytes'].quantile(0.25)
            Q3 = df['total_bytes'].quantile(0.75)
            IQR = Q3 - Q1
            
            lower_bound = Q1 -
